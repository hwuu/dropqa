"""文档解析器

支持多种文档格式：Markdown、Word、PowerPoint、Excel
"""

import re
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional


@dataclass
class ParsedNode:
    """解析后的节点"""
    node_type: str  # document, section, paragraph
    depth: int
    title: Optional[str] = None
    content: Optional[str] = None
    position: int = 0
    children: list["ParsedNode"] = field(default_factory=list)

    def __repr__(self) -> str:
        return f"ParsedNode(type={self.node_type}, depth={self.depth}, title={self.title!r})"


class MarkdownParser:
    """Markdown 解析器

    将 Markdown 文档解析为层级节点树。
    支持标题层级 (#, ##, ###, ...)
    """

    # 匹配 Markdown 标题
    HEADING_PATTERN = re.compile(r"^(#{1,6})\s+(.+)$")

    def parse(self, content: str, filename: str) -> ParsedNode:
        """解析 Markdown 内容

        Args:
            content: Markdown 文本内容
            filename: 文件名（用于根节点标题）

        Returns:
            解析后的文档节点树
        """
        lines = content.split("\n")

        # 创建根节点（document）
        root = ParsedNode(
            node_type="document",
            depth=0,
            title=filename,
        )

        # 解析行
        current_paragraph_lines: list[str] = []
        position = 0

        for line in lines:
            heading_match = self.HEADING_PATTERN.match(line)

            if heading_match:
                # 先保存之前积累的段落
                if current_paragraph_lines:
                    para_content = "\n".join(current_paragraph_lines).strip()
                    if para_content:
                        self._add_paragraph(root, para_content, position)
                        position += 1
                    current_paragraph_lines = []

                # 解析标题
                level = len(heading_match.group(1))  # # 数量
                title = heading_match.group(2).strip()

                section = ParsedNode(
                    node_type="section",
                    depth=level,
                    title=title,
                    position=position,
                )
                self._insert_node(root, section)
                position += 1
            else:
                # 收集段落行
                current_paragraph_lines.append(line)

        # 处理最后的段落
        if current_paragraph_lines:
            para_content = "\n".join(current_paragraph_lines).strip()
            if para_content:
                self._add_paragraph(root, para_content, position)

        return root

    def parse_file(self, file_path: Path) -> ParsedNode:
        """解析 Markdown 文件

        Args:
            file_path: 文件路径

        Returns:
            解析后的文档节点树
        """
        content = file_path.read_text(encoding="utf-8")
        return self.parse(content, file_path.name)

    def _insert_node(self, root: ParsedNode, node: ParsedNode) -> None:
        """将节点插入到正确的位置

        根据 depth 找到合适的父节点。
        """
        # 从根节点开始，找到合适的父节点
        parent = self._find_parent(root, node.depth)
        parent.children.append(node)

    def _find_parent(self, root: ParsedNode, target_depth: int) -> ParsedNode:
        """找到目标深度的父节点

        例如：目标 depth=2 (##)，应该找到 depth=1 (#) 的最后一个节点
        如果没有 depth=1，则挂在根节点下
        """
        if target_depth <= 1:
            return root

        # 从根节点递归查找
        return self._find_parent_recursive(root, target_depth)

    def _find_parent_recursive(self, node: ParsedNode, target_depth: int) -> ParsedNode:
        """递归查找父节点"""
        # 如果当前节点的子节点中没有 section，返回当前节点
        if not node.children:
            return node

        # 找最后一个 section 子节点
        last_section = None
        for child in reversed(node.children):
            if child.node_type == "section":
                last_section = child
                break

        if last_section is None:
            return node

        # 如果最后一个 section 的 depth < target_depth - 1，继续往下找
        if last_section.depth < target_depth - 1:
            return self._find_parent_recursive(last_section, target_depth)

        # 如果最后一个 section 的 depth == target_depth - 1，它就是父节点
        if last_section.depth == target_depth - 1:
            return last_section

        # 如果最后一个 section 的 depth >= target_depth，返回当前节点
        return node

    def _add_paragraph(self, root: ParsedNode, content: str, position: int) -> None:
        """添加段落节点"""
        # 找到当前最深的 section
        parent = self._find_last_section(root)
        para = ParsedNode(
            node_type="paragraph",
            depth=parent.depth + 1,
            content=content,
            position=position,
        )
        parent.children.append(para)

    def _find_last_section(self, node: ParsedNode) -> ParsedNode:
        """找到最后一个 section（用于添加段落）"""
        if not node.children:
            return node

        # 找最后一个 section 子节点
        for child in reversed(node.children):
            if child.node_type == "section":
                return self._find_last_section(child)

        return node


def flatten_nodes(root: ParsedNode, document_id: uuid.UUID, version: int = 1) -> list[dict]:
    """将节点树展平为列表（用于数据库插入）

    Args:
        root: 根节点
        document_id: 文档 ID
        version: 版本号

    Returns:
        展平后的节点列表，每个节点包含 id 和 parent_id
    """
    nodes = []
    _flatten_recursive(root, None, document_id, version, nodes)
    return nodes


def _flatten_recursive(
    node: ParsedNode,
    parent_id: Optional[uuid.UUID],
    document_id: uuid.UUID,
    version: int,
    nodes: list[dict],
) -> uuid.UUID:
    """递归展平节点"""
    node_id = uuid.uuid4()

    nodes.append({
        "id": node_id,
        "document_id": document_id,
        "version": version,
        "parent_id": parent_id,
        "node_type": node.node_type,
        "depth": node.depth,
        "title": node.title,
        "content": node.content,
        "position": node.position,
    })

    for child in node.children:
        _flatten_recursive(child, node_id, document_id, version, nodes)

    return node_id


# =============================================================================
# 多格式解析支持
# =============================================================================

# 支持的文件扩展名
SUPPORTED_EXTENSIONS = {".md", ".docx", ".pptx", ".xlsx"}


def parse_document(file_path: Path) -> ParsedNode:
    """统一解析入口（工厂模式）

    Args:
        file_path: 文件路径

    Returns:
        解析后的文档节点树

    Raises:
        ValueError: 不支持的文件格式
    """
    suffix = file_path.suffix.lower()

    if suffix == ".md":
        return MarkdownParser().parse_file(file_path)
    elif suffix == ".docx":
        return parse_docx(file_path)
    elif suffix == ".pptx":
        return parse_pptx(file_path)
    elif suffix == ".xlsx":
        return parse_xlsx(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {suffix}")


def parse_docx(file_path: Path) -> ParsedNode:
    """解析 Word 文档

    层级结构：
    - Heading 1 → section (depth=1)
    - Heading 2 → section (depth=2)
    - ...
    - 普通段落 → paragraph

    只有标题无内容的 section 会被过滤掉。

    Args:
        file_path: DOCX 文件路径

    Returns:
        解析后的文档节点树
    """
    from unstructured.partition.docx import partition_docx

    elements = partition_docx(str(file_path))

    # 创建根节点
    root = ParsedNode(
        node_type="document",
        depth=0,
        title=file_path.name,
    )

    position = 0
    current_sections: dict[int, ParsedNode] = {0: root}  # depth -> section

    for element in elements:
        element_type = type(element).__name__
        text = str(element).strip()

        if not text:
            continue

        if element_type == "Title":
            # unstructured 的 Title 对应标题，但不区分层级
            # 需要从 metadata 中获取层级信息
            metadata = element.metadata
            # 尝试从 category_depth 获取层级
            depth = getattr(metadata, "category_depth", None)
            if depth is None:
                depth = 1  # 默认为 1 级标题

            section = ParsedNode(
                node_type="section",
                depth=depth,
                title=text,
                position=position,
            )
            position += 1

            # 找到父节点
            parent_depth = depth - 1
            while parent_depth >= 0 and parent_depth not in current_sections:
                parent_depth -= 1
            parent = current_sections.get(parent_depth, root)
            parent.children.append(section)

            # 更新当前层级
            current_sections[depth] = section
            # 清除更深层级
            for d in list(current_sections.keys()):
                if d > depth:
                    del current_sections[d]

        elif element_type in ("NarrativeText", "ListItem", "Text"):
            # 普通段落
            # 找到当前最深的 section
            max_depth = max(current_sections.keys())
            parent = current_sections[max_depth]

            para = ParsedNode(
                node_type="paragraph",
                depth=parent.depth + 1,
                content=text,
                position=position,
            )
            position += 1
            parent.children.append(para)

    # 过滤只有标题无内容的 section
    _filter_empty_sections(root)

    return root


def parse_pptx(file_path: Path) -> ParsedNode:
    """解析 PowerPoint 文档

    层级结构：
    - 每个 slide → section (depth=1)
    - 每个文本框 → paragraph

    空白幻灯片会被跳过。

    Args:
        file_path: PPTX 文件路径

    Returns:
        解析后的文档节点树
    """
    from unstructured.partition.pptx import partition_pptx

    elements = partition_pptx(str(file_path))

    # 创建根节点
    root = ParsedNode(
        node_type="document",
        depth=0,
        title=file_path.name,
    )

    # 按页面分组
    slides: dict[int, list] = {}
    for element in elements:
        metadata = element.metadata
        page_number = getattr(metadata, "page_number", 1)
        if page_number not in slides:
            slides[page_number] = []
        slides[page_number].append(element)

    position = 0
    for page_num in sorted(slides.keys()):
        page_elements = slides[page_num]

        # 提取标题（第一个 Title 类型的元素）
        slide_title = None
        paragraphs = []

        for element in page_elements:
            element_type = type(element).__name__
            text = str(element).strip()

            if not text:
                continue

            if element_type == "Title" and slide_title is None:
                slide_title = text
            else:
                paragraphs.append(text)

        # 跳过空白幻灯片
        if not slide_title and not paragraphs:
            continue

        # 创建 slide section
        section = ParsedNode(
            node_type="section",
            depth=1,
            title=slide_title or f"第 {page_num} 页",
            position=position,
        )
        position += 1

        # 添加段落
        para_position = 0
        for para_text in paragraphs:
            para = ParsedNode(
                node_type="paragraph",
                depth=2,
                content=para_text,
                position=para_position,
            )
            para_position += 1
            section.children.append(para)

        root.children.append(section)

    return root


def parse_xlsx(file_path: Path) -> ParsedNode:
    """解析 Excel 文档

    层级结构：
    - 每个 sheet → section (depth=1)
    - 每行 → section (depth=2)，第一行作为表头

    空行会被跳过。

    Args:
        file_path: XLSX 文件路径

    Returns:
        解析后的文档节点树
    """
    from openpyxl import load_workbook

    wb = load_workbook(str(file_path), read_only=True, data_only=True)

    # 创建根节点
    root = ParsedNode(
        node_type="document",
        depth=0,
        title=file_path.name,
    )

    sheet_position = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 创建 sheet section
        sheet_section = ParsedNode(
            node_type="section",
            depth=1,
            title=sheet_name,
            position=sheet_position,
        )
        sheet_position += 1

        # 读取所有行
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # 第一行作为表头
        headers = rows[0]
        header_names = [str(h) if h is not None else f"列{i+1}" for i, h in enumerate(headers)]

        # 从第二行开始处理数据
        row_position = 0
        for row_idx, row in enumerate(rows[1:], start=2):
            # 跳过空行
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            # 构建行内容：字段名: 值
            content_parts = []
            for i, cell in enumerate(row):
                if cell is not None and str(cell).strip():
                    field_name = header_names[i] if i < len(header_names) else f"列{i+1}"
                    content_parts.append(f"{field_name}: {cell}")

            if not content_parts:
                continue

            content = ", ".join(content_parts)

            row_section = ParsedNode(
                node_type="section",
                depth=2,
                title=f"第{row_idx}行",
                content=content,
                position=row_position,
            )
            row_position += 1
            sheet_section.children.append(row_section)

        # 只添加非空的 sheet
        if sheet_section.children:
            root.children.append(sheet_section)

    wb.close()
    return root


def _filter_empty_sections(node: ParsedNode) -> bool:
    """过滤只有标题无内容的 section

    Args:
        node: 节点

    Returns:
        True 表示该节点应该保留，False 表示应该删除
    """
    if node.node_type == "paragraph":
        # 段落节点：有内容则保留
        return bool(node.content and node.content.strip())

    if node.node_type == "document":
        # 根节点：递归处理子节点
        node.children = [child for child in node.children if _filter_empty_sections(child)]
        return True

    if node.node_type == "section":
        # section 节点：递归处理子节点
        node.children = [child for child in node.children if _filter_empty_sections(child)]
        # 如果有子节点则保留，否则删除
        return len(node.children) > 0

    return True
